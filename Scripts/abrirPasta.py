### Importar libs ###
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import pandas as pd
import shutil
import os
import sys
import re
from selenium import webdriver
from time import sleep
from selenium.webdriver.chrome.options import Options
sys.path.append(r'C:\Users\MatheusPereira\OneDrive - Pontte\Área de Trabalho\ProjCredito\automacaoCreditoFinanciamento\Scripts')       
from anexosCliente import listarAnexos
from baixarAnexo import baixarEMover
from datetime import datetime
import openpyxl as xl
from openpyxl.worksheet.datavalidation import DataValidation
import glob
from baixarBase import downloadRelatorio, atualizarBase, finalAtualizacao
sleep(14)

### Funçoes ###

def deParaimovel(varImovel):
    if varImovel == "Casa Padrão" or varImovel == "Casa de Condomínio":
        varImovel = "Casa"
    return varImovel
def deParaOps(varOps):
    if varOps == "Pessoa Física":
        varOps = "PF"
    elif varOps == "Pessoa Jurídica":
        varOps = "PJ"
    return varOps
def converterDatas(date):
    try:
        date_obj = datetime.strptime(str(date), '%Y-%m-%d')
        date = date_obj.strftime("%d/%m/%Y")
    except:
        pass
    return date


### Baixar e Atualizar Base ###

pathDestino = r'G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Casos_Analise_Bacen\Base_Casos.xlsx'
downloadRelatorio('https://app.pipefy.com/pipes/301487926/reports_v2/300284180')
atualizarBase(pathDestino)
finalAtualizacao()

# Criar Data Frame da Base Atualizada
pathBase = r'G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Casos_Analise_Bacen\Base_Casos.xlsx'
df = pd.read_excel(pathBase)
df[['Data de Nascimento - 2o Pagador(a)', 'Qual a sua data de nascimento?']].value_counts()

final = len(df)
for i in range(0,final):
    
    codigoPipefy = df.iloc[i]['Código']
    linkPipefy = "https://app.pipefy.com/open-cards/" + str(codigoPipefy)
    nomeCompleto = df.iloc[i]['Qual é o seu nome completo?']
    tipoOps = df.iloc[i]['Tipo de Operação']
    tipoOps = deParaOps(tipoOps)
    cpf = df.iloc[i]['E o seu CPF?']
    dataNascimento = df.iloc[i]['Qual a sua data de nascimento?']
    dataNascimento = converterDatas(dataNascimento)
    comporRenda = df.iloc[i]['Deseja compor renda com alguém?']
    nomeComporRenda = df.iloc[i]['Nome Completo - 2o Pagador(a)']
    cpfComporRenda = df.iloc[i]['CPF - 2o Pagador(a)']
    dataNascimentoComporRenda = df.iloc[i]['Data de Nascimento - 2o Pagador(a)']
    dataNascimentoComporRenda = converterDatas(dataNascimentoComporRenda)
    valorImovel = df.iloc[i]['Qual o valor total do imóvel que você quer comprar?']
    valorLiquido = df.iloc[i]['Quanto do valor do imóvel você quer financiar?']
    qtdParcelas = df.iloc[i]['Selecione a quantidade de parcelas do seu financiamento']
    qtdParcelas = re.sub('[^0-9]', '', qtdParcelas)
    carencia = df.iloc[i]['Você precisa de uma carência antes de começar a pagar suas parcelas?']
    carencia = re.sub('[^0-9]', '', carencia)
    #new
    tipoImovel = df.iloc[i]['Qual é o tipo de Imóvel?']
    tipoImovel = deParaimovel(tipoImovel)
    eSocio = df.iloc[i]['Você é sócio(a) de alguma empresa?']
    razaoSocial = df.iloc[i]['Razão Social Empresa']
    cnpjEmpresa = df.iloc[i]['CNPJ Empresa']
    areaImovel = df.iloc[i]['Qual é a área útil do imóvel?']
    enderecoImovel = df.iloc[i]['Qual o endereço do imóvel de interesse?']
    itbi = enderecoImovel = df.iloc[i]['Deseja incluir ITBI e Registro no financiamento?']

    origem = r"G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Simulação\1_Template"
    destino = rf"G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Simulação\{nomeCompleto} - {codigoPipefy}" 

    

    if not os.path.exists(destino):
        shutil.copytree(origem, destino)
        dict = listarAnexos(linkPipefy, destino)
        if nomeCompleto is not None:
            renomearCPF1 = rf'G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Simulação\{nomeCompleto} - {codigoPipefy}\Crédito\Documentos pessoais\CPF1'
            renomeadoCPF1 = rf'G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Simulação\{nomeCompleto} - {codigoPipefy}\Crédito\Documentos pessoais\{nomeCompleto} - {cpf}'
            os.rename(renomearCPF1,renomeadoCPF1)
            if comporRenda == "Sim":
                renomearCPF2 = rf'G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Simulação\{nomeCompleto} - {codigoPipefy}\Crédito\Documentos pessoais\CPF2'
                renomeadoCPF2 = rf'G:\Drives compartilhados\Pontte\Operações\Projetos\Automação de Crédito\Simulação\{nomeCompleto} - {codigoPipefy}\Crédito\Documentos pessoais\{nomeComporRenda} - {cpfComporRenda}'
                os.rename(renomearCPF2,renomeadoCPF2)
        
        baixarEMover(destino, dict, nomeCompleto, cpf, nomeComporRenda, cpfComporRenda)

        #Manipulação do simulador do Excel
        try:
            pathSpreadSheet = glob.glob(destino+"\*.xlsm") #Pegar Caminho do Simulador dentro da pasta do cliente
            pathSpreadSheet = pathSpreadSheet[0] #Transformar de Lista ---> STR
            wb = xl.load_workbook(pathSpreadSheet, read_only=False, keep_vba=True) #Carregar a planilha
            #Abrir Abas
            ws = wb['Resumo']
            ws2 = wb['Avaliacao']
            #Escrever nas planilhas
            ws["C8"].value = int(qtdParcelas)
            if carencia == "Não informado" or carencia == 'Não tenho interesse' or carencia == '':
                ws["C10"].value = 0
            else:    
                ws["C10"].value = carencia
            ws["C11"].value = valorLiquido
            ws["C12"].value = valorImovel
            ws["C13"].value = valorImovel
            ws["C14"].value = itbi
            ws["C15"].value = tipoOps
            ws["C16"].value = tipoImovel
            #Nomes
            ws["C19"].value = nomeCompleto
            ws["C20"].value = nomeComporRenda
            ws["C21"].value = razaoSocial
            #dataNascimento
            ws["C24"].value = dataNascimento
            ws["C25"].value = dataNascimentoComporRenda
            #Link
            ws["C29"].value = linkPipefy
            #CPF e CNPJ
            ws["C30"].value = cpf
            ws["C31"].value = cpfComporRenda
            ws["C32"].value = cnpjEmpresa
            
            #Avaliacao
            ws2["F9"].value = areaImovel
            
            #Salvar e Fechar
            wb.save(pathSpreadSheet)
            wb.close()
            print('Finalizado!')
        except:
            print('Erro ao preencher Excel!')
