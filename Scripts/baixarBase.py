from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from time import sleep
import pandas as pd
import openpyxl as xl
import os
from datetime import date
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) 

################# INICIO INFORMAÇÕES INICIAIS ###################

#Caminho para iniciar ChromeDriver
options = webdriver.ChromeOptions() 
options.add_experimental_option('excludeSwitches', ['enable-logging'])
#options.add_argument("--headless")
driver = webdriver.Chrome(options=options,executable_path=r'G:\Drives compartilhados\Pontte\Operações\Automações\contaQI\Driver\chromedriver.exe')

#Gerar dia de hoje no Formato DD-MM-AAAA
data_atual = date.today()
data_atual = data_atual.strftime('%d-%m-%Y')

#Pegar nome do Usuário
usuario = os.environ.get('USERNAME')

#Ajustar Downloads Antes de Rodar o Código
try: 
    try:
        os.remove(pathOrigem)
    except:
        pass
    try:
        os.remove(rf"C:\Users\{usuario}\Downloads\relatório_automação_crédito_{data_atual} (1).xlsx")
    except:
        pass
    try:
        os.remove(rf"C:\Users\{usuario}\Downloads\relatório_automação_crédito_{data_atual} (2).xlsx")
    except:
        pass
except:
    pass

#################  FINAL INFORMAÇÕES INICIAIS ###################


################# INCIO FUNÇÕES DO SCRIPT ###################

def fazerLogin():
    #Site que irá abrir
    driver.get('https://app.pipefy.com/')

    sleep(5)

    #### Inicio Processo de Login ####

    #Identifica e retorna os elementos
    print('Inicio do processo de login')
    inserirEmail_element = driver.find_element_by_name('username')
    inseirSenha_element =  driver.find_element_by_name('password')
    entrar_element =  driver.find_element_by_name('submit')

    #InserirValores
    inserirEmail_element.send_keys('matheus.pereira@pontte.com.br')
    inseirSenha_element.send_keys('Pontte2000')
    print('login realizado')

    #Clicar
    entrar_element.click()
    sleep(2)

def downloadRelatorio(linkRelatorio):
    print('Inicio do Processo de Download')
    driver.get(linkRelatorio)
    sleep(5)

    #### Geração de Relatório ####

    driver.find_element_by_xpath('//*[@id="pipe_placeholder"]/div[2]/div/div[1]/div[2]/div[1]/div[1]/div[3]/div/div/button').click()
    sleep(5)
    status = None
    i = 1
    while status == None and i <= 20:
        try:
            clickDownload = "/html/body/div["+str(i)+"]/div/div[2]/footer/button[2]"
            sleep(1)
            driver.find_element_by_xpath(clickDownload).click()
            status = 'Feito'
            print(status)
            sleep(5)
        except:
            i = i+1
    sleep(7)    
    print('Final do Processo de Download')

def atualizarBase(pathDestino):
    #Caminho Relatório Baixado
    pathOrigem = rf"C:\Users\{usuario}\Downloads\relatório_automação_crédito_{data_atual}.xlsx"
    sleep(5)
    base = xl.load_workbook(pathOrigem, data_only=True)
    origem = base["Report"]
    destino = xl.load_workbook(pathDestino, data_only=False)#, keep_vba=True#)
    dt = destino["Base"]
    def cp(r,c,i,j):
        valor = origem.cell(row=r, column=c)
        dt.cell(row=i, column=j).value = valor.value

    def clearExcel(i,j):
        dt.cell(row=i, column=j).value = ""

    print('Inicio do Processo')

    for c in range (1,40):
        for r in range(1,50):
            i = r
            j = c
            clearExcel(i,j)
            cp(r, c, i, j)
            
    print('Final do processo')

    destino.save(str(pathDestino))
    driver.close()

#################  FINAL FUNÇÕES DO SCRIPT ###################

fazerLogin()
pathOrigem = rf"C:\Users\{usuario}\Downloads\relatório_automação_crédito_{data_atual}.xlsx"

def finalAtualizacao():
    #Excluir Arquivo da Pasta Download
    try: 
        os.remove(pathOrigem)
    except:
        pass